"""Excel 文档生成 Helper 函数
封装 openpyxl 常用操作，内置专业配色方案
"""

import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# 专业配色方案
if HAS_OPENPYXL:
    THEME = {
        "header_fill": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
        "alt_row_fill": PatternFill(start_color="EDF2F9", end_color="EDF2F9", fill_type="solid"),
        "header_font": Font(name="微软雅黑", bold=True, color="1F4E79", size=11),
        "title_font": Font(name="微软雅黑", bold=True, color="1F4E79", size=16),
        "normal_font": Font(name="微软雅黑", size=11),
        "border": Border(
            left=Side(style='thin', color='B4C6E7'),
            right=Side(style='thin', color='B4C6E7'),
            top=Side(style='thin', color='B4C6E7'),
            bottom=Side(style='thin', color='B4C6E7'),
        ),
        "center_align": Alignment(horizontal='center', vertical='center'),
    }
else:
    THEME = {}


def create_excel_doc(title=None, author=""):
    """创建一个预配置好专业样式的 Excel 工作簿对象

    Args:
        title: 工作簿标题（可选）
        author: 文档作者

    Returns:
        Workbook: 预配置好的 openpyxl Workbook 对象

    示例:
        wb = create_excel_doc(title="销售数据", author="张三")
        ws = wb.active
        ws.title = "销售数据"
        ws.append(["产品", "销量", "金额"])
        save_excel_doc(wb, "销售数据.xlsx")
    """
    wb = Workbook()
    # 移除默认工作表，让用户自行添加
    # wb.remove(wb.active)

    if title:
        wb.properties.title = title
    if author:
        wb.properties.creator = author

    return wb


def save_excel_doc(wb, filename, working_dir=""):
    """保存 Excel 工作簿到工作目录

    Args:
        wb: openpyxl Workbook 对象
        filename: 文件名（如 "数据.xlsx"）
        working_dir: 工作目录路径

    Returns:
        str: 保存的文件绝对路径
    """
    output_path = os.path.join(working_dir, filename) if working_dir else filename
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)

    return output_path
