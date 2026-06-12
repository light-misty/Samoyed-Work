"""Excel 文档处理器
基于 openpyxl 实现 Excel 文档的读取、转换、分析
精简版：仅支持 read/convert/analyze 操作
"""

import os
import csv
import io
import html
import logging

from openpyxl import load_workbook


class ExcelHandler:
    """Excel (.xlsx) 文档处理器（精简版，仅支持 read/convert/analyze）"""

    logger = logging.getLogger(__name__)

    def read(self, params: dict) -> dict:
        """读取 Excel 文档

        params:
            path: 文件路径
            sheet: 工作表名称（可选，默认读取所有）
            range: 读取范围（可选，如 "A1:D10"）
        """
        path = params.get("path", "")
        sheet_name = params.get("sheet", None)
        read_range = params.get("range", None)
        if not path:
            self.logger.error("read: 缺少文件路径")
            return {"error": "缺少文件路径"}
        if not os.path.exists(path):
            raise FileNotFoundError(path)

        self.logger.info("read: 开始读取 Excel 文档, path=%s", path)

        wb = load_workbook(path, data_only=False)
        result = {"sheets": {}}

        if sheet_name:
            sheets_to_read = [sheet_name]
        else:
            sheets_to_read = wb.sheetnames

        for name in sheets_to_read:
            if name not in wb.sheetnames:
                continue
            ws = wb[name]

            if read_range:
                rows = []
                for row in ws[read_range]:
                    rows.append([cell.value for cell in row])
            else:
                rows = []
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
                    row_data = []
                    for cell in row:
                        val = cell.value
                        if val is None and cell.data_type == "f":
                            val = cell.internal_value
                        row_data.append(val)
                    rows.append(row_data)

            result["sheets"][name] = {
                "data": rows,
                "row_count": ws.max_row,
                "col_count": ws.max_column,
            }

        result["sheet_names"] = wb.sheetnames
        self.logger.info("read: Excel 文档读取完成, path=%s, 工作表数=%d", path, len(result["sheets"]))
        return result

    def convert(self, params: dict) -> dict:
        """格式转换

        params:
            path: 源文件路径
            output_path: 输出文件路径（可选）
            format: 目标格式（csv, pdf, html, txt）
            sheet: 工作表名称（可选，默认转换所有）
        """
        path = params.get("path", "")
        output_path = params.get("output_path", "")
        target_format = params.get("format", "").lower()
        sheet_name = params.get("sheet", None)

        if not path:
            self.logger.error("convert: 缺少源文件路径")
            return {"error": "缺少源文件路径"}
        if not os.path.exists(path):
            raise FileNotFoundError(path)
        if target_format not in ("csv", "pdf", "html", "txt"):
            self.logger.error("convert: 不支持的目标格式: %s", target_format)
            return {"error": f"不支持的目标格式: {target_format}，支持的格式: csv, pdf, html, txt"}

        self.logger.info("convert: 开始格式转换, path=%s, format=%s", path, target_format)

        wb = load_workbook(path, data_only=True)

        # 确定要转换的工作表
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                self.logger.error("convert: 工作表不存在: %s", sheet_name)
                return {"error": f"工作表不存在: {sheet_name}"}
            sheets_to_convert = [sheet_name]
        else:
            sheets_to_convert = wb.sheetnames

        # 读取所有工作表数据
        all_sheets_data = {}
        for name in sheets_to_convert:
            ws = wb[name]
            rows = []
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
                rows.append([cell if cell is not None else "" for cell in row])
            all_sheets_data[name] = rows

        # PDF 为二进制格式，必须写入文件
        if target_format == "pdf":
            if not output_path:
                base, _ = os.path.splitext(path)
                output_path = base + ".pdf"
            self._convert_to_pdf(all_sheets_data, output_path)
            self.logger.info("convert: 格式转换完成, output_path=%s, format=%s", output_path, target_format)
            return {
                "path": output_path,
                "format": target_format,
                "message": f"已转换为 {target_format} 格式",
            }

        # 根据目标格式生成文本内容
        if target_format == "csv":
            content = self._convert_to_csv(all_sheets_data)
        elif target_format == "html":
            content = self._convert_to_html(all_sheets_data)
        elif target_format == "txt":
            content = self._convert_to_txt(all_sheets_data)

        # 写入输出文件或返回内容
        if output_path:
            os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
            with open(output_path, "w", encoding="utf-8") as f:
                f.write(content)
            self.logger.info("convert: 格式转换完成, output_path=%s, format=%s", output_path, target_format)
            return {
                "path": output_path,
                "format": target_format,
                "message": f"已转换为 {target_format} 格式",
            }
        else:
            return {
                "content": content,
                "format": target_format,
            }

    def analyze(self, params: dict) -> dict:
        """分析 Excel 文档"""
        path = params.get("path", "")
        if not path:
            self.logger.error("analyze: 缺少文件路径")
            return {"error": "缺少文件路径"}
        if not os.path.exists(path):
            raise FileNotFoundError(path)

        self.logger.info("analyze: 开始分析 Excel 文档, path=%s", path)

        wb = load_workbook(path, data_only=True)
        sheets_info = []
        for name in wb.sheetnames:
            ws = wb[name]
            sheets_info.append({
                "name": name,
                "rows": ws.max_row,
                "cols": ws.max_column,
            })

        self.logger.info("analyze: Excel 文档分析完成, path=%s, 工作表数=%d", path, len(wb.sheetnames))
        return {
            "file_size": os.path.getsize(path),
            "sheet_count": len(wb.sheetnames),
            "sheets": sheets_info,
        }

    # ------------------------------------------------------------------ #
    #  格式转换辅助方法
    # ------------------------------------------------------------------ #

    def _convert_to_csv(self, all_sheets_data: dict) -> str:
        """将工作表数据转换为 CSV 格式"""
        parts = []
        for sheet_name, rows in all_sheets_data.items():
            if len(all_sheets_data) > 1:
                parts.append(f"# 工作表: {sheet_name}")
            output = io.StringIO()
            writer = csv.writer(output)
            for row in rows:
                writer.writerow(row)
            parts.append(output.getvalue().rstrip("\r\n"))

        return "\n".join(parts)

    def _convert_to_pdf(self, all_sheets_data: dict, output_path: str):
        """将工作表数据转换为 PDF 格式（使用 reportlab 渲染表格）"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        except ImportError:
            self.logger.error("convert: reportlab 未安装，无法转换为 PDF")
            raise RuntimeError("reportlab 未安装，无法转换为 PDF")

        from handlers.font_utils import register_chinese_font
        font_name = register_chinese_font()

        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

        doc = SimpleDocTemplate(output_path, pagesize=A4)
        styles = getSampleStyleSheet()

        sheet_title_style = ParagraphStyle(
            "SheetTitle",
            parent=styles["Normal"],
            fontName=font_name,
            fontSize=14,
            spaceAfter=10,
            spaceBefore=20,
        )

        elements = []
        for sheet_name, rows in all_sheets_data.items():
            elements.append(Paragraph(html.escape(sheet_name), sheet_title_style))
            elements.append(Spacer(1, 0.5 * cm))

            if rows:
                table_data = [[str(cell) for cell in row] for row in rows]

                table = Table(table_data)
                table.setStyle(TableStyle([
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("FONTNAME", (0, 0), (-1, -1), font_name),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.9, 0.9, 0.9)),
                    ("FONTNAME", (0, 0), (-1, 0), font_name),
                    ("FONTSIZE", (0, 0), (-1, 0), 9),
                ]))
                elements.append(table)
                elements.append(Spacer(1, 1 * cm))

        doc.build(elements)

    def _convert_to_html(self, all_sheets_data: dict) -> str:
        """将工作表数据转换为 HTML 表格"""
        parts = [
            "<!DOCTYPE html>",
            "<html>",
            "<head>",
            '<meta charset="utf-8">',
            "<title>Excel 转换</title>",
            "<style>",
            "body { font-family: sans-serif; margin: 20px; }",
            "table { border-collapse: collapse; margin-bottom: 20px; width: 100%; }",
            "th, td { border: 1px solid #ccc; padding: 6px 10px; text-align: left; }",
            "th { background-color: #f0f0f0; font-weight: bold; }",
            "h2 { margin-top: 30px; color: #333; }",
            "</style>",
            "</head>",
            "<body>",
        ]

        for sheet_name, rows in all_sheets_data.items():
            parts.append(f"<h2>{html.escape(sheet_name)}</h2>")
            parts.append("<table>")
            for i, row in enumerate(rows):
                tag = "th" if i == 0 else "td"
                parts.append("<tr>")
                for cell in row:
                    parts.append(f"<{tag}>{html.escape(str(cell))}</{tag}>")
                parts.append("</tr>")
            parts.append("</table>")

        parts.extend(["</body>", "</html>"])
        return "\n".join(parts)

    def _convert_to_txt(self, all_sheets_data: dict) -> str:
        """将工作表数据转换为纯文本（制表符分隔）"""
        parts = []
        for sheet_name, rows in all_sheets_data.items():
            if len(all_sheets_data) > 1:
                parts.append(f"=== {sheet_name} ===")
            for row in rows:
                parts.append("\t".join(str(cell) for cell in row))
            parts.append("")

        return "\n".join(parts)
